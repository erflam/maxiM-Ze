import sys
import os
import time

if getattr(sys, 'frozen', False):
    sys.path.insert(0, sys._MEIPASS)
else:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import shutil
import multiprocessing
from multiprocessing import cpu_count
from pathlib import Path

from Config import Config
from FileUtils import FileUtils
from FileReader import process_file_checkpoint1
from EICBuilder import process_file_checkpoint2
from Resolving import process_file_checkpoint3, count_peaks_per_file_summary
from PixelMapping import process_file_checkpoint4
from Slicing import process_file_checkpoint5
from Coelution import run_group_coelution
from CoelutionSliced import process_file_coelution_sliced
from Clustering import process_file_cluster_peaks
from Recluster import process_file_recluster
from Visualization import process_visualizations
from ExportExcel import process_export_excel
from LibraryMatching import process_library_match

def init_worker():
    import matplotlib
    matplotlib.use('Agg')
    import os
    os.environ['KALEIDO_DISABLE'] = '1'
    os.environ['PLOTLY_RENDERER'] = 'json'

class Pipeline:
    def __init__(self, import_json_path: str | Path | None = None):
        self.config = Config()
        self.file_paths = FileUtils.get_file_paths()
        self.timings = []  # [{step, group, elapsed_s}]

        # Time mass grouping — runs before run() so must be recorded here
        _mg_start = time.time()
        Config.initialize_mass_groups(
            self.file_paths,
            import_json_path=import_json_path,
        )
        self._record("Mass Grouping / Detection", None, time.time() - _mg_start)

        self.file_colors = {
            os.path.splitext(os.path.basename(fp))[0]: FileUtils.random_dark_hex_color()
            for fp in self.file_paths
        }

    # ------------------------------------------------------------------
    # Timing helpers
    # ------------------------------------------------------------------
    def _record(self, step_name: str, group_name: str | None, elapsed: float):
        self.timings.append({
            "Step": step_name,
            "Group": group_name if group_name is not None else "—",
            "Elapsed (s)": round(elapsed, 2),
        })

    def _export_timings(self):
        """Write self.timings to pipeline_timings.xlsx in the analysis output folder."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timings"

        # Header row
        headers = ["Step", "Group", "Elapsed (s)"]
        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="2F5496")
        center = Alignment(horizontal="center")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Separate step rows from the PIPELINE TOTAL sentinel
        step_rows = [e for e in self.timings if e["Step"] != "PIPELINE TOTAL"]
        pipeline_total = next((e for e in self.timings if e["Step"] == "PIPELINE TOTAL"), None)

        arial = Font(name="Arial")
        bold_arial = Font(name="Arial", bold=True)

        for row_idx, entry in enumerate(step_rows, 2):
            ws.cell(row=row_idx, column=1, value=entry["Step"]).font = arial
            ws.cell(row=row_idx, column=2, value=entry["Group"]).font = arial
            ws.cell(row=row_idx, column=3, value=entry["Elapsed (s)"]).font = arial

        # STEPS TOTAL — sums only the individual step rows (no double-count)
        sum_row = len(step_rows) + 2
        ws.cell(row=sum_row, column=1, value="STEPS TOTAL").font = bold_arial
        ws.cell(row=sum_row, column=3, value=f"=SUM(C2:C{sum_row - 1})").font = bold_arial

        # PIPELINE TOTAL (wall-clock) sits below the sum for comparison
        if pipeline_total:
            pt_row = sum_row + 1
            ws.cell(row=pt_row, column=1, value="PIPELINE TOTAL (wall clock)").font = bold_arial
            ws.cell(row=pt_row, column=2, value=pipeline_total["Group"]).font = arial
            ws.cell(row=pt_row, column=3, value=pipeline_total["Elapsed (s)"]).font = bold_arial

        # Column widths
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        out_path = Config._analysis_output_root() / "pipeline_timings.xlsx"
        wb.save(out_path)
        print(f"Timing report saved → {out_path.resolve()}")
        return out_path

    # ------------------------------------------------------------------

    def _delete_group_outputs(self, group_name):
        Config.set_mass_group(group_name)
        dirs = Config.setup_directories()
        for key, dir_path in dirs.items():
            if dir_path and os.path.exists(dir_path):
                shutil.rmtree(dir_path)
                print(f"Deleted [{group_name}] {key}: {dir_path}")

    def clean_run(self):
        print("\nCLEAN RUN: deleting all previous outputs...\n")
        for group_name in Config.get_group_names_to_run():
            Config.set_mass_group(group_name)
            dirs = Config.setup_directories()
            for dir_path in dirs.values():
                if dir_path and os.path.exists(dir_path):
                    shutil.rmtree(dir_path)
                    print(f"Deleted {dir_path}")
        print("\nAll outputs deleted. Starting fresh run...\n")
        self.run()

    def run_group_checkpoint1(self, dirs, group_name):
        files_to_process = [fp for fp in self.file_paths if fp and os.path.exists(fp)]
        if not files_to_process:
            print("No valid files found.")
            return

        t0 = time.time()
        args = [(fp, dirs, group_name) for fp in files_to_process]
        ctx = multiprocessing.get_context('spawn')
        with ctx.Pool(processes=max(1, cpu_count() - 1), initializer=init_worker) as pool:
            results = pool.starmap(process_file_checkpoint1, args)
        for r in results:
            print(r)
        elapsed = time.time() - t0
        self._record("Checkpoint 1 – File Reading", group_name, elapsed)
        print(f'Checkpoint 1 completed in {elapsed:.2f} seconds!')

    def run_group_checkpoint2(self, dirs, group_name):
        files_to_process = [fp for fp in self.file_paths if fp and os.path.exists(fp)]
        if not files_to_process:
            print("No valid files found.")
            return

        t0 = time.time()
        args = [(fp, dirs, self.file_colors, group_name) for fp in files_to_process]
        ctx = multiprocessing.get_context('spawn')
        with ctx.Pool(processes=max(1, cpu_count() - 1), initializer=init_worker) as pool:
            results = pool.starmap(process_file_checkpoint2, args)
        for r in results:
            print(r)
        elapsed = time.time() - t0
        self._record("Checkpoint 2 – EIC Builder", group_name, elapsed)
        print(f'Checkpoint 2 completed in {elapsed:.2f} seconds!')

    def run_group_checkpoint3(self, dirs, group_name):
        files_to_process = [fp for fp in self.file_paths if fp and os.path.exists(fp)]
        if not files_to_process:
            print("No valid files found.")
            return

        t0 = time.time()
        args = [(fp, dirs, group_name) for fp in files_to_process]
        ctx = multiprocessing.get_context('spawn')
        with ctx.Pool(processes=max(1, cpu_count() - 1), initializer=init_worker) as pool:
            results = pool.starmap(process_file_checkpoint3, args)
        for r in results:
            print(r)
        print(count_peaks_per_file_summary(dirs, group_name))
        elapsed = time.time() - t0
        self._record("Checkpoint 3 – Resolving", group_name, elapsed)
        print(f'Checkpoint 3 completed in {elapsed:.2f} seconds!')

    def run_group_checkpoint4(self, dirs, group_name):
        png_dir = dirs['png']
        group_tag = str(group_name).replace(" ", "")
        pngs = [
            os.path.join(png_dir, f)
            for f in os.listdir(png_dir)
            if f.startswith("EIC_") and f.endswith(f"_{group_tag}.png")
        ]
        if not pngs:
            print(f"[!] No PNGs found for {group_name} in {png_dir}")
            try:
                for x in sorted(os.listdir(png_dir))[:10]:
                    print("    -", x)
            except Exception:
                pass
            return

        t0 = time.time()
        args = [(png, dirs, group_name) for png in pngs]
        ctx = multiprocessing.get_context('spawn')
        with ctx.Pool(processes=max(1, cpu_count() - 1), initializer=init_worker) as pool:
            results = pool.starmap(process_file_checkpoint4, args)
        for r in results:
            print(r)
        elapsed = time.time() - t0
        self._record("Checkpoint 4 – Pixel Mapping", group_name, elapsed)
        print(f"Checkpoint 4 (Pixel Mapping) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint5(self, dirs, group_name):
        t0 = time.time()
        process_file_checkpoint5(self, dirs, group_name)
        elapsed = time.time() - t0
        self._record("Checkpoint 5 – Slicing", group_name, elapsed)
        print(f"Checkpoint 5 (Slicing based on Pixel Mapping) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint6(self, dirs, group_name):
        t0 = time.time()
        run_group_coelution(dirs=dirs, group_name=group_name)
        elapsed = time.time() - t0
        self._record("Checkpoint 6 – Coelution Slices", group_name, elapsed)
        print(f"Checkpoint 6 (Coelution slices added to Directory) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint7(self, dirs, group_name):
        t0 = time.time()
        msg = process_file_coelution_sliced(dirs, group_name)
        print(msg)
        elapsed = time.time() - t0
        self._record("Checkpoint 7 – Coelution Valley Reslicing", group_name, elapsed)
        print(f"Checkpoint 7 (Coelution valley reslicing) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint8(self, dirs, group_name):
        t0 = time.time()
        msg = process_file_cluster_peaks(dirs, group_name)
        print(msg)
        elapsed = time.time() - t0
        self._record("Checkpoint 8 – Peak Clustering/Alignment", group_name, elapsed)
        print(f"Checkpoint 8 (Peak clustering/alignment) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint9(self, dirs: dict, group_name: str) -> None:
        t0 = time.time()
        msg = process_file_recluster(dirs, group_name)
        print(msg)
        elapsed = time.time() - t0
        self._record("Checkpoint 9 – RT+Mass Recluster", group_name, elapsed)
        print(f"Checkpoint 9 (Post-clustering RT+mass recluster) completed in {elapsed:.2f} seconds!")

    def run_group_checkpoint10(self, dirs, group_name):
        t0 = time.time()
        self.dirs = dirs
        msg = process_visualizations(self, group_name)
        print(msg)
        elapsed = time.time() - t0
        self._record("Checkpoint 10 – Visual QC Composites", group_name, elapsed)
        print(f"Checkpoint 10 (Visual QC composites) completed in {elapsed:.2f} seconds!")

    def run_final_checkpoint_excel(self) -> Path:
        t0 = time.time()
        excel_path = process_export_excel(Config)
        print(f"Excel export complete → {excel_path}")
        elapsed = time.time() - t0
        self._record("Final – Excel Export", None, elapsed)
        print(f"Final Checkpoint (Excel export) completed in {elapsed:.2f} seconds!")
        return excel_path

    def run_final_checkpoint_library_match(self, excel_path):
        t0 = time.time()
        msg = process_library_match(Config, excel_path)
        print(msg)
        elapsed = time.time() - t0
        self._record("Final – Library Match", None, elapsed)
        print(f"Final Checkpoint (Library Match) completed in {elapsed:.2f} seconds!")

    def run(self):
        total_start = time.time()

        for group_name in Config.get_group_names_to_run():
            print(f"Running group: {group_name}")
            Config.set_mass_group(group_name)
            dirs = Config.setup_directories()
            group_tag = str(group_name).replace(" ", "")

            self.run_group_checkpoint1(dirs, group_name)
            self.run_group_checkpoint2(dirs, group_name)

            png_dir = dirs['png']
            built_pngs = [
                f for f in os.listdir(png_dir)
                if f.startswith("EIC_") and f.endswith(f"_{group_tag}.png")
            ] if os.path.exists(png_dir) else []

            if not built_pngs:
                print(f"[!] No EIC PNGs built for group '{group_name}' "
                      f"(all masses likely noisy or below noise threshold). Skipping group.")
                continue

            self.run_group_checkpoint3(dirs, group_name)
            self.run_group_checkpoint4(dirs, group_name)
            self.run_group_checkpoint5(dirs, group_name)
            self.run_group_checkpoint6(dirs, group_name)
            self.run_group_checkpoint7(dirs, group_name)
            self.run_group_checkpoint8(dirs, group_name)
            self.run_group_checkpoint9(dirs, group_name)
            self.run_group_checkpoint10(dirs, group_name)

        excel_path = self.run_final_checkpoint_excel()
        lib_file = getattr(Config, 'LIB_FILE', None)
        if lib_file and os.path.exists(lib_file):
            self.run_final_checkpoint_library_match(excel_path)
        else:
            print("[–] No library file provided or file not found — skipping library matching.")

        total_elapsed = time.time() - total_start
        self._record("PIPELINE TOTAL", None, total_elapsed)
        print(f"\nTotal pipeline time: {total_elapsed:.2f} seconds")

        self._export_timings()

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    multiprocessing.set_start_method("spawn", force=True)
    pipeline = Pipeline()
    pipeline.clean_run()