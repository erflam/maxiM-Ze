import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _group_sort_key(name: str):
    m = re.search(r"(\d+)", str(name))
    return int(m.group(1)) if m else str(name)


def _append_sheet_rows(
    src_xlsx: Path,
    src_sheet_name: str,
    dest_ws,
    start_row: int,
    include_header: bool = True,
) -> int:
    """
    Appends rows from src_sheet_name into dest_ws starting at start_row (0-based).
    If include_header is False, skips the first row of the source sheet.
    Returns number of rows written.
    """
    src_wb = openpyxl.load_workbook(src_xlsx, data_only=True)
    try:
        if src_sheet_name not in src_wb.sheetnames:
            return 0

        src_ws = src_wb[src_sheet_name]
        rows_written = 0

        for row_idx, src_row in enumerate(src_ws.iter_rows()):
            is_header = (row_idx == 0)
            if is_header and not include_header:
                continue

            dest_row_idx = start_row + rows_written + 1  # openpyxl is 1-based
            for src_cell in src_row:
                dest_cell = dest_ws.cell(
                    row=dest_row_idx,
                    column=src_cell.column,
                    value=src_cell.value,
                )
                if src_cell.font and src_cell.font.bold:
                    dest_cell.font = Font(bold=True)

            rows_written += 1

        return rows_written
    finally:
        src_wb.close()


def _count_data_rows(src_xlsx: Path, sheet_name: str) -> int:
    """
    Counts "data rows" in a sheet, assuming first row is header.
    Returns max(0, max_row - 1).
    """
    src_wb = openpyxl.load_workbook(src_xlsx, data_only=True, read_only=True)
    try:
        if sheet_name not in src_wb.sheetnames:
            return 0
        ws = src_wb[sheet_name]
        max_row = ws.max_row or 0
        return max(0, max_row - 1)
    finally:
        src_wb.close()


def _autosize_columns(ws):
    """
    Simple autosize based on string length of cell values.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def export_all_group_summaries_to_excel(Config) -> Path:
    output_root = Path(Config.BASE_DIR) / Path(Config.OUTPUT_ROOT) / Config.ANALYSIS_FOLDER
    excel_path = output_root / f"MassSelectionSummary_{Config.ANALYSIS_FOLDER}.xlsx"
    output_root.mkdir(parents=True, exist_ok=True)

    print(f"\n[Excel Export] Saving summary workbook to {excel_path}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sorted_groups = sorted(Config.MASS_GROUPS.keys(), key=_group_sort_key)

    # ------------------------------------------------------------------
    # SHEET 1: All_Groups_Summary (concatenated "Summary" sheets)
    # ------------------------------------------------------------------
    master_ws = wb.create_sheet(title="All_Groups_Summary")
    master_row = 0
    wrote_header = False

    # ------------------------------------------------------------------
    # SHEET 2: Groups_With_Unclustered_Peaks
    # ------------------------------------------------------------------
    unclustered_ws = wb.create_sheet(title="Groups_With_Unclustered_Peaks")
    unclustered_ws.append(["Group", "Unclustered_Peak_Count", "Group_Summary_File"])
    for cell in unclustered_ws[1]:
        cell.font = Font(bold=True)

    groups_with_unclustered = 0

    for group_name in sorted_groups:
        cluster_dir = output_root / str(group_name) / "Clustering"
        group_summary_xlsx = cluster_dir / f"Group_Summary_{group_name}.xlsx"

        if not group_summary_xlsx.exists():
            continue

        # Load once to check sheet existence
        src_wb = openpyxl.load_workbook(group_summary_xlsx, read_only=True, data_only=True)
        try:
            sheet_names = set(src_wb.sheetnames)
        finally:
            src_wb.close()

        # Append Summary into master
        if "Summary" in sheet_names:
            rows_written = _append_sheet_rows(
                group_summary_xlsx,
                "Summary",
                master_ws,
                start_row=master_row,
                include_header=(not wrote_header),
            )
            if rows_written > 0:
                master_row += rows_written
                wrote_header = True
                print(f"[✔] Master sheet ← {group_name} Summary ({rows_written} rows)")
        else:
            # no Summary sheet; skip
            continue

        # Track which groups have Unclustered peaks
        if "Unclustered" in sheet_names:
            peak_count = _count_data_rows(group_summary_xlsx, "Unclustered")
            if peak_count > 0:  # <-- Only add if actual data rows exist
                unclustered_ws.append([
                    str(group_name),
                    int(peak_count),
                    str(group_summary_xlsx)
                ])
                groups_with_unclustered += 1
                print(f"    ↳ Unclustered found for {group_name} ({peak_count} rows)")

    _autosize_columns(master_ws)
    _autosize_columns(unclustered_ws)

    wb.save(excel_path)
    print(
        f"\n[✔] Excel export complete → {excel_path}\n"
        f"[!] Groups with unclustered peaks: {groups_with_unclustered}\n"
    )
    return excel_path


def process_export_excel(Config) -> str:
    excel_path = export_all_group_summaries_to_excel(Config)
    return f"Excel export complete → {excel_path}"