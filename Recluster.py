import time
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class Config:
    BASE_DIR        = Path(".")
    OUTPUT_ROOT     = "Output"
    ANALYSIS_FOLDER = "Analysis"
    CURRENT_GROUP   = "Group1"

    @classmethod
    def clustering_dir(cls) -> Path:
        return cls.BASE_DIR / cls.OUTPUT_ROOT / cls.ANALYSIS_FOLDER / cls.CURRENT_GROUP / "Clustering"

    @classmethod
    def pixel_dir(cls) -> Path:
        return cls.BASE_DIR / cls.OUTPUT_ROOT / cls.ANALYSIS_FOLDER / cls.CURRENT_GROUP / "Pixel CSVs"

    @classmethod
    def peak_patches_dir(cls) -> Path:
        # Used only for filename normalization/verification (optional)
        return cls.BASE_DIR / cls.OUTPUT_ROOT / cls.ANALYSIS_FOLDER / cls.CURRENT_GROUP / "Peak Patches"

RT_TOLERANCE = 0.08

def _sample_name_from_col(col: str) -> str:
    for suffix in ("_height", "_area"):
        if col.endswith(suffix):
            return col[: -len(suffix)]
    return col

def _sample_name_from_peak_id(peak_id: str) -> str:
    idx = peak_id.find("_mass")
    if idx != -1:
        return peak_id[:idx]
    return peak_id

def _height_col(sample: str) -> str:
    return f"{sample}_height"

def _area_col(sample: str) -> str:
    return f"{sample}_area"

def _to_float(x):
    if x is None:
        return None
    if isinstance(x, (float, int, np.floating, np.integer)):
        if pd.isna(x):
            return None
        return float(x)
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return None
    try:
        return float(s)
    except Exception:
        return None

def _to_int(x):
    if x is None:
        return None
    if isinstance(x, (int, np.integer)):
        return int(x)
    if isinstance(x, (float, np.floating)):
        if pd.isna(x):
            return None
        try:
            return int(round(float(x)))
        except Exception:
            return None
    s = str(x).strip()
    if s == "" or s.lower() in {"nan", "none"}:
        return None
    try:
        return int(round(float(s)))
    except Exception:
        return None

def _png_name_from_peak_id(peak_id: str) -> str:
    peak_id = str(peak_id).strip()
    if not peak_id:
        return peak_id
    return peak_id if peak_id.lower().endswith(".png") else f"{peak_id}.png"

def _build_patch_index(patch_dir: Path) -> dict[str, str]:
    index: dict[str, str] = {}
    if patch_dir and patch_dir.exists():
        for p in patch_dir.glob("*.png"):
            index[p.name.lower()] = p.name
    return index

def _load_cluster_patch(path: Path, drop_empty_clusters: bool = True) -> list[dict]:
    if not path.exists():
        return []

    df = pd.read_csv(path, header=0)
    if df.empty or len(df.columns) < 2:
        return []

    # First column is the row-label column (empty header or similar)
    label_col = df.columns[0]
    cluster_cols = [c for c in df.columns if str(c).startswith("Cluster ")]

    label_series = df[label_col].fillna("").astype(str).tolist()

    clusters: list[dict] = []
    for col in cluster_cols:
        values = df[col].tolist()
        row_map = dict(zip(label_series, values))

        patches: list[str] = []
        for label, val in zip(label_series, values):
            # PNG rows are stored where label == "" (per your writer)
            if label == "" and pd.notna(val) and str(val).strip():
                patches.append(str(val).strip())

        mz = _to_float(row_map.get("m/z"))
        iso = _to_int(row_map.get("Isomer_position"))
        art = _to_float(row_map.get("Aligned_rt_apex"))
        pc  = _to_int(row_map.get("peak count"))

        # If peak count is missing/incorrect, recompute from patches
        if pc is None:
            pc = len(patches)

        cluster = {
            "col_name": str(col),
            "m/z": mz,
            "Isomer_position": iso,
            "Aligned_rt_apex": art,
            "peak count": pc,
            "patches": patches,
        }

        if drop_empty_clusters and len(patches) == 0:
            # Drop clusters that contain no patch filenames.
            continue

        clusters.append(cluster)

    return clusters

def _clusters_to_df(clusters: list[dict]) -> pd.DataFrame:
    if not clusters:
        return pd.DataFrame()

    max_patches = max(len(c["patches"]) for c in clusters) if clusters else 0
    meta_labels = ["m/z", "Isomer_position", "Aligned_rt_apex", "peak count"]
    index_col = meta_labels + [""] * max_patches

    data: dict[str, list] = {"": index_col}
    for cluster in clusters:
        col_values = [
            cluster.get("m/z"),
            cluster.get("Isomer_position"),
            cluster.get("Aligned_rt_apex"),
            cluster.get("peak count"),
        ]
        patches = list(cluster.get("patches") or [])
        patches = patches + [""] * (max_patches - len(patches))
        col_values.extend(patches)
        data[str(cluster.get("col_name"))] = col_values

    return pd.DataFrame(data)

def recluster_group(group: str, config: type = Config) -> None:
    clustering_dir = config.clustering_dir()
    pixel_dir      = config.pixel_dir()

    alignment_path   = clustering_dir / f"alignment_summary_group_{group}.csv"
    unclustered_path = clustering_dir / f"unclustered_peaks_group_{group}.csv"
    patch_path       = clustering_dir / f"cluster_patch_{group}.csv"
    output_path      = clustering_dir / f"Group_Summary_{group}.xlsx"

    patch_index = _build_patch_index(getattr(config, "peak_patches_dir", lambda: None)())

    alignment   = pd.read_csv(alignment_path)
    unclustered = pd.read_csv(unclustered_path)
    clusters    = _load_cluster_patch(patch_path, drop_empty_clusters=True)

    # Normalize core fields
    if "m/z" in alignment.columns:
        alignment["m/z"] = alignment["m/z"].apply(_to_float)
    if "Isomer_position" in alignment.columns:
        alignment["Isomer_position"] = alignment["Isomer_position"].apply(_to_int)
    if "m/z" in unclustered.columns:
        unclustered["m/z"] = unclustered["m/z"].apply(_to_float)
    if "RT_apex" in unclustered.columns:
        unclustered["RT_apex"] = pd.to_numeric(unclustered["RT_apex"], errors="coerce")
    if "height" in unclustered.columns:
        unclustered["height"] = pd.to_numeric(unclustered["height"], errors="coerce")
    if "area" in unclustered.columns:
        unclustered["area"] = pd.to_numeric(unclustered["area"], errors="coerce")
    if "peak_id" in unclustered.columns:
        unclustered["peak_id"] = unclustered["peak_id"].astype(str)
    if "peak_id" in unclustered.columns:
        unclustered["_sample"] = unclustered["peak_id"].apply(_sample_name_from_peak_id)
    else:
        unclustered["_sample"] = ""

    height_cols = [c for c in alignment.columns if str(c).endswith("_height")]
    area_cols   = [c for c in alignment.columns if str(c).endswith("_area")]
    all_samples = [_sample_name_from_col(c) for c in height_cols]

    alignment["Recluster"] = False

    newly_filled: set[tuple[int, str]] = set()
    matched_peak_ids: set[str] = set()
    peaks_used_count = 0
    cluster_append_failures = 0

    has_unclustered = (
        not unclustered.empty
        and "m/z" in unclustered.columns
        and "RT_apex" in unclustered.columns
        and "peak_id" in unclustered.columns
    )

    if has_unclustered:
        unclustered_grouped = {}
        for (mz_val, sample), grp in unclustered.groupby(["m/z", "_sample"], dropna=False):
            unclustered_grouped[(mz_val, sample)] = grp

        for row_idx, row in alignment.iterrows():
            aligned_rt = _to_float(row.get("Aligned_rt_apex"))
            mz         = _to_float(row.get("m/z"))
            iso_pos    = _to_int(row.get("Isomer_position", None))

            if mz is None or aligned_rt is None:
                continue

            for sample in all_samples:
                h_col = _height_col(sample)
                a_col = _area_col(sample)

                height_val = row.get(h_col, np.nan)
                if pd.notna(height_val) and float(height_val) != 0.0:
                    continue

                grp = unclustered_grouped.get((mz, sample))
                if grp is None or grp.empty:
                    continue

                # Candidate filter: RT tolerance + NOT already matched
                candidates = grp[
                    (grp["RT_apex"].notna()) &
                    ((grp["RT_apex"] - aligned_rt).abs() <= RT_TOLERANCE) &
                    (~grp["peak_id"].isin(matched_peak_ids))
                ].copy()

                if candidates.empty:
                    continue

                candidates["_rt_diff"] = (candidates["RT_apex"] - aligned_rt).abs()
                best = candidates.loc[candidates["_rt_diff"].idxmin()]

                # Fill alignment cells
                alignment.at[row_idx, h_col] = best.get("height", np.nan)
                newly_filled.add((row_idx, h_col))

                if a_col in alignment.columns:
                    alignment.at[row_idx, a_col] = best.get("area", np.nan)
                    newly_filled.add((row_idx, a_col))

                alignment.at[row_idx, "Recluster"] = True

                peak_id = str(best["peak_id"])
                matched_peak_ids.add(peak_id)
                peaks_used_count += 1

                png_name = _png_name_from_peak_id(peak_id)
                # Normalize to actual filename in directory when possible
                png_name = patch_index.get(png_name.lower(), png_name)

                found_cluster = False
                for cluster in clusters:
                    if cluster.get("m/z") == mz and cluster.get("Isomer_position") == iso_pos:
                        found_cluster = True
                        if png_name not in cluster["patches"]:
                            cluster["patches"].append(png_name)
                        cluster["peak count"] = len(cluster["patches"])
                        break

                if not found_cluster:
                    cluster_append_failures += 1

    for row_idx, row in alignment.iterrows():
        if not bool(row.get("Recluster", False)):
            continue

        heights = [row.get(_height_col(s), np.nan) for s in all_samples]
        new_count = sum(1 for h in heights if pd.notna(h) and float(h) != 0.0)
        if "peak count" in alignment.columns:
            alignment.at[row_idx, "peak count"] = int(new_count)

        mz_val = _to_float(row.get("m/z"))
        aligned_rt = _to_float(row.get("Aligned_rt_apex"))
        if mz_val is None or aligned_rt is None:
            continue

        rt_values: list[float] = []
        for sample in all_samples:
            h_col = _height_col(sample)
            hv = row.get(h_col, np.nan)
            if pd.notna(hv) and float(hv) != 0.0:
                pixel_file = pixel_dir / f"{sample}_peaks_pix_{group}.csv"
                if not pixel_file.exists():
                    continue
                try:
                    pix = pd.read_csv(pixel_file)
                except Exception:
                    continue

                if "m/z" not in pix.columns or "RT_apex" not in pix.columns:
                    continue

                pix["m/z"] = pix["m/z"].apply(_to_float)
                pix["RT_apex"] = pd.to_numeric(pix["RT_apex"], errors="coerce")

                match = pix[
                    (pix["m/z"] == mz_val) &
                    (pix["RT_apex"].notna()) &
                    ((pix["RT_apex"] - aligned_rt).abs() <= (RT_TOLERANCE + 0.05))
                ]

                if not match.empty:
                    best_idx = (match["RT_apex"] - aligned_rt).abs().idxmin()
                    best_rt = match.loc[best_idx, "RT_apex"]
                    if pd.notna(best_rt):
                        rt_values.append(float(best_rt))

        if rt_values:
            alignment.at[row_idx, "Aligned_rt_apex"] = float(np.mean(rt_values))

    if "peak_id" in unclustered.columns:
        still_unclustered = unclustered[~unclustered["peak_id"].isin(matched_peak_ids)].copy()
        if "_sample" in still_unclustered.columns:
            still_unclustered.drop(columns=["_sample"], inplace=True)
    else:
        still_unclustered = unclustered.copy()

    for c in clusters:
        c["peak count"] = len(c.get("patches") or [])

    _write_excel(
        output_path   = output_path,
        alignment     = alignment,
        unclustered   = still_unclustered,
        clusters      = clusters,
        newly_filled  = newly_filled,
        height_cols   = height_cols,
        area_cols     = area_cols,
    )

    print(f"Done. Output written to: {output_path}")
    print(f"  Rows reclustered         : {int(alignment['Recluster'].sum())}")
    print(f"  Peaks matched (unique)   : {len(matched_peak_ids)}")
    print(f"  Peak assignments used    : {peaks_used_count}")
    print(f"  Still unclustered        : {len(still_unclustered)}")
    print(f"  Clusters loaded          : {len(clusters)} (empty clusters dropped)")
    if cluster_append_failures:
        print(f"  WARN: matched peaks not appended to any cluster: {cluster_append_failures}")

def _write_excel(
    output_path:  Path,
    alignment:    pd.DataFrame,
    unclustered:  pd.DataFrame,
    clusters:     list[dict],
    newly_filled: set[tuple[int, str]],
    height_cols:  list[str],
    area_cols:    list[str],
) -> None:
    wb = Workbook()

    bold_font   = Font(name="Arial", bold=True)
    normal_font = Font(name="Arial", bold=False)
    header_font = Font(name="Arial", bold=True)

    ws1 = wb.active
    ws1.title = "Summary"

    cols = list(alignment.columns)
    if "Recluster" in cols:
        cols.remove("Recluster")
        pc_idx = cols.index("peak count") if "peak count" in cols else len(cols) - 1
        cols.insert(pc_idx + 1, "Recluster")
    alignment = alignment[cols]

    for col_idx, col_name in enumerate(alignment.columns, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    col_name_to_excel_col = {name: i + 1 for i, name in enumerate(alignment.columns)}

    for df_row_idx, (_, row) in enumerate(alignment.iterrows()):
        excel_row    = df_row_idx + 2
        orig_df_idx  = alignment.index[df_row_idx]

        for col_name in alignment.columns:
            value = row[col_name]
            if isinstance(value, (np.bool_,)):
                value = bool(value)
            elif isinstance(value, (np.integer,)):
                value = int(value)
            elif isinstance(value, (np.floating,)):
                value = float(value) if not np.isnan(value) else None

            excel_col = col_name_to_excel_col[col_name]
            cell = ws1.cell(row=excel_row, column=excel_col, value=value)
            cell.font = bold_font if (orig_df_idx, col_name) in newly_filled else normal_font

    for col_cells in ws1.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws1.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

    ws2 = wb.create_sheet(title="Unclustered")

    if not unclustered.empty:
        for col_idx, col_name in enumerate(unclustered.columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, (_, row) in enumerate(unclustered.iterrows(), start=2):
            for col_idx, col_name in enumerate(unclustered.columns, start=1):
                value = row[col_name]
                if isinstance(value, (np.bool_,)):
                    value = bool(value)
                elif isinstance(value, (np.integer,)):
                    value = int(value)
                elif isinstance(value, (np.floating,)):
                    value = float(value) if not np.isnan(value) else None
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font

        for col_cells in ws2.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws2.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)
    else:
        ws2.cell(row=1, column=1, value="No unclustered peaks remaining.")

    ws3 = wb.create_sheet(title="Cluster PNGs")

    df_patch = _clusters_to_df(clusters)

    if not df_patch.empty:
        meta_rows = {"m/z", "Isomer_position", "Aligned_rt_apex", "peak count"}

        for col_idx, col_name in enumerate(df_patch.columns, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx, (_, row) in enumerate(df_patch.iterrows(), start=2):
            label = str(row.iloc[0]) if row.iloc[0] else ""
            is_meta = label in meta_rows

            for col_idx, col_name in enumerate(df_patch.columns, start=1):
                value = row[col_name]
                if pd.isna(value) or value == "":
                    value = None
                elif isinstance(value, (np.integer,)):
                    value = int(value)
                elif isinstance(value, (np.floating,)):
                    value = float(value) if not np.isnan(value) else None

                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.font = bold_font if is_meta else normal_font

        for col_cells in ws3.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws3.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
    else:
        ws3.cell(row=1, column=1, value="No cluster patch data available.")

    wb.save(output_path)

def process_file_recluster(dirs: dict, group_name: str) -> str:
    class _RuntimeConfig(Config):
        @classmethod
        def clustering_dir(cls) -> Path:
            return Path(dirs["clustering"])

        @classmethod
        def pixel_dir(cls) -> Path:
            return Path(dirs["pixel"])

        @classmethod
        def peak_patches_dir(cls) -> Path:
            # Optional, only used for PNG filename normalization
            return Path(dirs.get("peak_patches", "")) if dirs.get("peak_patches") else super().peak_patches_dir()

    recluster_group(group=group_name, config=_RuntimeConfig)
    output_path = Path(dirs["clustering"]) / f"Group_Summary_{group_name}.xlsx"
    return f"Reclustering complete for {group_name}. Output: {output_path}"

def run_group_checkpoint9(self, dirs: dict, group_name: str) -> None:
    start_time = time.time()
    msg = process_file_recluster(dirs, group_name)
    print(msg)
    elapsed = time.time() - start_time
    print(f"Checkpoint 9 (Post-clustering RT+mass recluster) completed in {elapsed:.2f} seconds!")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Recluster unclustered peaks into alignment summary.")
    parser.add_argument("group", help="Group identifier, e.g. Group1")
    parser.add_argument("--base-dir",        default=".",        help="Project base directory")
    parser.add_argument("--output-root",     default="Output",   help="Output root folder name")
    parser.add_argument("--analysis-folder", default="Analysis", help="Analysis folder name")
    args = parser.parse_args()

    Config.BASE_DIR        = Path(args.base_dir)
    Config.OUTPUT_ROOT     = args.output_root
    Config.ANALYSIS_FOLDER = args.analysis_folder
    Config.CURRENT_GROUP   = args.group

    recluster_group(group=args.group, config=Config)